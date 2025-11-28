"""
Result export module - Handles exporting scraped results to CSV/Excel/JSON formats
"""
import pandas as pd
import logging
from datetime import datetime
from pathlib import Path
from utils import create_output_dir
from config import OUTPUT_FORMAT

logger = logging.getLogger(__name__)

# Export scraped results to CSV/Excel/JSON format.
# Args:
#     results: List of case dictionaries to export
#     output_dir: Output directory path. If None or empty, uses default from create_output_dir()
# Returns:
#     None
def export_results(results, output_dir):

    if not results:
        print("No results to export")
        return
    
    # Create DataFrame from results
    df = pd.DataFrame(results)
    
    # Define proper column titles mapping
    column_titles = {
        "attorney_name": "Attorney",
        "attorney_first_name": "Attorney First Name",
        "attorney_last_name": "Attorney Last Name",
        "case_number": "Case Number",
        "file_date": "File Date",
        "judicial_officer": "Judicial Officer",
        "case_status": "Case Status",
        "case_type": "Case Type",
        "charge_description": "Charge Description",
        "bond_amount": "Bond Amount",
        "disposition": "Disposition",
        "sentencing_info": "Sentencing"
    }
    
    # Rename columns to proper titles
    df = df.rename(columns=column_titles)
    
    # Reorder columns to ensure logical order (Attorney info first)
    column_order = [
        "Attorney",
        "Case Number",
        "File Date",
        "Judicial Officer",
        "Case Status",
        "Case Type",
        "Charge Description",
        "Bond Amount",
        "Disposition",
        "Sentencing"
    ]
    
    # Only include columns that exist in the dataframe
    column_order = [col for col in column_order if col in df.columns]
    df = df[column_order]
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Convert output_dir to Path up front for uniform handling
    if output_dir is not None:
        output_dir = Path(output_dir)
    
    # Use provided output_dir if available and non-empty, otherwise use default
    if output_dir is None or not str(output_dir).strip():
        output_dir = create_output_dir()
    else:
        # Ensure directory exists
        output_dir.mkdir(parents=True, exist_ok=True)
    
    if OUTPUT_FORMAT.lower() == "csv":
        export_csv(df, output_dir, timestamp)
    elif OUTPUT_FORMAT.lower() == "excel":
        export_excel(df, output_dir, timestamp)
    elif OUTPUT_FORMAT.lower() == "json":
        export_json(df, output_dir, timestamp)
    else:
        print(f"Unknown output format: {OUTPUT_FORMAT}. Defaulting to Excel.")
        export_excel(df, output_dir, timestamp)
    
    # Print summary statistics
    print_export_summary(df, results)

# Export results to CSV format
def export_csv(df, output_dir, timestamp):
   
    filename = output_dir / f"cases_{timestamp}.csv"
    df.to_csv(filename, index=False)
    print(f"Results exported to: {filename}")

# Export results to JSON format
def export_json(df, output_dir, timestamp):

    filename = output_dir / f"cases_{timestamp}.json"
    df.to_json(filename, orient="records", indent=2)
    print(f"Results exported to: {filename}")

# Sanitize name for Excel sheet name (31 char limit, no invalid chars)
def sanitize_sheet_name(name, max_length=31):
    """
    Sanitize a name to be a valid Excel sheet name.
    Excel sheet names cannot contain: / \ ? * [ ]
    Cannot start or end with apostrophe
    Max length is 31 characters
    """
    # Replace invalid characters
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, '-')
    
    # Remove leading/trailing spaces and apostrophes
    name = name.strip().strip("'")
    
    # Truncate to max length
    if len(name) > max_length:
        name = name[:max_length].rstrip()
    
    # Ensure name is not empty
    if not name:
        name = "Sheet"
    
    return name

# Export results to Excel format with formatting and multiple sheets
def export_excel(df, output_dir, timestamp):
    filename = output_dir / f"cases_{timestamp}.xlsx"
    
    try:
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write "All Cases" sheet with all results
        df.to_excel(writer, sheet_name='All Cases', index=False)
        
        # Create mapping of ACTUAL sheet names (as created by Excel) to original attorney names
        sheet_to_attorney = {}
        used_sheet_names = set(['All Cases'])  # Track used names to avoid duplicates
        
        # Create separate sheets for each attorney
        if "Attorney" in df.columns:
            attorneys = df["Attorney"].unique()
            for attorney in attorneys:
                if pd.notna(attorney) and attorney != "UNKNOWN":
                    # Filter cases for this attorney
                    attorney_df = df[df["Attorney"] == attorney].copy()
                    
                    # Create base sanitized sheet name
                    base_sheet_name = sanitize_sheet_name(str(attorney))
                    
                    # Ensure uniqueness by appending number if needed
                    sheet_name = base_sheet_name
                    counter = 1
                    while sheet_name in used_sheet_names:
                        # Truncate base name further to make room for counter suffix
                        available_length = 31 - len(str(counter)) - 1  # -1 for space or dash
                        base_name = base_sheet_name[:available_length].rstrip()
                        sheet_name = f"{base_name}-{counter}"
                        counter += 1
                    
                    used_sheet_names.add(sheet_name)
                    
                    # Record number of sheets before adding this one
                    workbook = writer.book
                    sheet_count_before = len(workbook.sheetnames)
                    
                    # Write sheet
                    attorney_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the ACTUAL sheet name as created by Excel/openpyxl
                    # (Excel might sanitize it further, so we use the actual name)
                    sheet_count_after = len(workbook.sheetnames)
                    if sheet_count_after > sheet_count_before:
                        # New sheet was added - get the last sheet name (which is the one we just added)
                        actual_sheet_name = workbook.sheetnames[-1]
                    else:
                        # Fallback: use the calculated name (shouldn't happen, but be safe)
                        actual_sheet_name = sheet_name
                    
                    # Store mapping using actual sheet name that Excel created
                    sheet_to_attorney[actual_sheet_name] = attorney
                    
                    # Update used names set with actual sheet name (in case Excel modified it)
                    if actual_sheet_name != sheet_name:
                        used_sheet_names.add(actual_sheet_name)
        
        # Format all sheets (pass mapping to resolve sheet names correctly)
        format_excel_sheets(writer, df, sheet_to_attorney)
    
    print(f"Results exported to: {filename}")
    if "Attorney" in df.columns:
        attorneys = df["Attorney"].unique()
        attorney_count = len([a for a in attorneys if pd.notna(a) and a != "UNKNOWN"])
        print(f"Created {attorney_count + 1} sheets: 'All Cases' + {attorney_count} attorney sheet(s)")

# Apply formatting to Excel sheets (column widths, header styling)
# Args:
#     writer: ExcelWriter object
#     df: Original DataFrame with all data
#     sheet_to_attorney: Mapping of sheet names to original attorney names
def format_excel_sheets(writer, df, sheet_to_attorney=None):

    from openpyxl.styles import Font, PatternFill, Alignment
    
    workbook = writer.book
    
    # Format all sheets
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Get the dataframe for this sheet
        if sheet_name == 'All Cases':
            sheet_df = df
        else:
            # Use mapping to find original attorney name from actual sheet name
            if sheet_to_attorney and sheet_name in sheet_to_attorney:
                attorney_name = sheet_to_attorney[sheet_name]
                sheet_df = df[df["Attorney"] == attorney_name] if "Attorney" in df.columns else df
            else:
                # Fallback: try to match by finding attorney name that sanitizes to this sheet name
                # This handles edge cases where mapping might not be perfect
                if "Attorney" in df.columns:
                    # Try to reverse-engineer the attorney name from the sheet name
                    matched = False
                    for attorney in df["Attorney"].unique():
                        if pd.notna(attorney) and attorney != "UNKNOWN":
                            sanitized = sanitize_sheet_name(str(attorney))
                            if sanitized == sheet_name or sheet_name.startswith(sanitized.rstrip('-')):
                                sheet_df = df[df["Attorney"] == attorney]
                                matched = True
                                break
                    if not matched:
                        # Last resort: empty dataframe
                        sheet_df = df.iloc[0:0].copy() if len(df) > 0 else df
                else:
                    sheet_df = df
        
        # Auto-adjust column widths
        for idx, col in enumerate(sheet_df.columns, 1):
            try:
                # Get max length of content in column
                max_length = max(
                    sheet_df[col].astype(str).apply(len).max(),
                    len(col)
                )
                # Set column width (add some padding)
                worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = min(max_length + 2, 50)
            except (AttributeError, IndexError, KeyError, ValueError) as e:
                # Log exception details for debugging
                logger.warning(f"Error calculating column width for column '{col}' (index {idx}): {e}")
            except (ValueError, TypeError):
                # Fallback width if calculation fails
                worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = 15        # Format header row (bold, background color)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

# Print summary statistics after export
def print_export_summary(df, results):
    print(f"\nTotal cases found: {len(results)}")
    
    # Show breakdown by attorney if available
    if "Attorney" in df.columns:
        print("\nCases by Attorney:")
        attorney_counts = df["Attorney"].value_counts()
        for attorney, count in attorney_counts.items():
            print(f"  - {attorney}: {count} case(s)")
    
    print("\nColumn titles:")
    for col in df.columns:
        print(f"  - {col}")
    print("\nFirst few results:")
    print(df.head().to_string())

